import logging
from pathlib import Path
import openpyxl
from urllib.parse import urlparse

logger = logging.getLogger(__name__)


class ValidationError(Exception):
    """Raised when input validation fails."""

    pass


class WorkbookValidator:
    """Validates workbook files and sources for security and format compliance."""

    # Configuration constants
    MAX_FILE_SIZE_MB = 50  # 50MB limit
    MAX_SHEETS = 20  # Reasonable sheet limit
    ALLOWED_EXTENSIONS = {".xlsx", ".xls"}

    def validate_workbook_source(
        self, workbook_source: str, is_local_file: bool
    ) -> None:
        """
        Validate workbook source (URL or file path).

        Args:
            workbook_source: URL or file path to validate
            is_local_file: Whether source is local file or URL

        Raises:
            ValidationError: If validation fails
        """
        if not workbook_source or not workbook_source.strip():
            raise ValidationError("Workbook source cannot be empty")

        if is_local_file:
            self._validate_local_file(workbook_source)
        else:
            self._validate_url(workbook_source)

    def validate_workbook_file(self, file_path: Path) -> None:
        """
        Validate actual workbook file.

        Args:
            file_path: Path to workbook file

        Raises:
            ValidationError: If validation fails
        """
        if not file_path.exists():
            raise ValidationError(f"Workbook file not found: {file_path}")

        # Check file extension
        if file_path.suffix.lower() not in self.ALLOWED_EXTENSIONS:
            raise ValidationError(
                f"Invalid file format. Allowed formats: {', '.join(self.ALLOWED_EXTENSIONS)}"
            )

        # Check file size
        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > self.MAX_FILE_SIZE_MB:
            raise ValidationError(
                f"File too large ({file_size_mb:.1f}MB). Maximum allowed: {self.MAX_FILE_SIZE_MB}MB"
            )

        # Validate Excel file integrity
        self._validate_excel_format(file_path)

    def _validate_local_file(self, file_path: str) -> None:
        """Validate local file path."""
        path = Path(file_path)

        # Basic path traversal protection
        try:
            path.resolve(strict=True)
        except (OSError, RuntimeError):
            raise ValidationError("Invalid file path")

        # Check if file exists and is readable
        if not path.exists():
            raise ValidationError(f"File not found: {file_path}")

        if not path.is_file():
            raise ValidationError(f"Path is not a file: {file_path}")

    def _validate_url(self, url: str) -> None:
        """Validate URL format."""
        try:
            parsed = urlparse(url)
            if not parsed.scheme or not parsed.netloc:
                raise ValidationError("Invalid URL format")

            if parsed.scheme not in ("http", "https"):
                raise ValidationError("Only HTTP and HTTPS URLs are allowed")

        except Exception as e:
            raise ValidationError(f"Invalid URL: {str(e)}")

    def _validate_excel_format(self, file_path: Path) -> None:
        """Validate Excel file can be opened and basic structure."""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)

            # Check sheet count
            if len(workbook.sheetnames) > self.MAX_SHEETS:
                workbook.close()  # Clean up before raising
                raise ValidationError(
                    f"Too many sheets ({len(workbook.sheetnames)}). Maximum allowed: {self.MAX_SHEETS}"
                )

            # Check if workbook has at least one sheet with data
            has_data = False
            for sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]

                    # Handle None values from max_row/max_column
                    max_row = sheet.max_row or 0
                    max_column = sheet.max_column or 0

                    # Check if sheet has meaningful data (more than just headers)
                    if max_row > 1 and max_column > 0:
                        has_data = True
                        break

                    # Alternative check: look for any non-None values
                    if max_row <= 1:
                        # Check first few cells for data
                        for row in sheet.iter_rows(
                            min_row=1, max_row=min(5, max_row or 1)
                        ):
                            for cell in row:
                                if cell.value is not None and str(cell.value).strip():
                                    has_data = True
                                    break
                            if has_data:
                                break

                except Exception as e:
                    logger.warning(f"Error checking sheet '{sheet_name}': {e}")
                    continue  # Skip problematic sheets, check others

            workbook.close()  # Always close workbook

            if not has_data:
                raise ValidationError(
                    "Workbook appears to be empty or contains only empty sheets"
                )

        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValidationError("File is not a valid Excel workbook")
        except ValidationError:
            raise  # Re-raise validation errors
        except Exception as e:
            raise ValidationError(f"Cannot read Excel file: {str(e)}")


def validate_analysis_request(
    workbook_source: str, is_local_file: bool, instruction: str
) -> None:
    """
    Validate complete analysis request.

    Args:
        workbook_source: URL or file path
        is_local_file: Whether source is local file
        instruction: Analysis instruction

    Raises:
        ValidationError: If validation fails
    """
    validator = WorkbookValidator()

    # Validate workbook source
    validator.validate_workbook_source(workbook_source, is_local_file)

    # Validate instruction
    if not instruction or not instruction.strip():
        raise ValidationError("Analysis instruction cannot be empty")

    if len(instruction) > 10000:
        raise ValidationError("Analysis instruction too long (max 10,000 characters)")
