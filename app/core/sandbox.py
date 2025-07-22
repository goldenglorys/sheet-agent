import code
import io
import sys
from pathlib import Path
import ast 
import os

from app.utils.common import SandboxResponse
from app.utils.enumeration import EXEC_CODE 
import logging

logger = logging.getLogger(__name__)

TRIM_SHEET_CODE = """
from openpyxl.utils import get_column_letter

def trim_sheet(ws):
    max_data_row = 0
    max_data_col = 0

    # Scan all cells to find the maximum row/column that contains data
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                if cell.row > max_data_row:
                    max_data_row = cell.row
                if cell.column > max_data_col:
                    max_data_col = cell.column

    # If the sheet is entirely empty, nothing to trim
    if max_data_row == 0 or max_data_col == 0:
        return

    # Delete rows below max_data_row
    # Note: delete from the bottom up to avoid reindexing issues
    for row_idx in range(ws.max_row, max_data_row, -1):
        ws.delete_rows(row_idx)

    # Delete columns to the right of max_data_col
    for col_idx in range(ws.max_column, max_data_col, -1):
        ws.delete_cols(col_idx)

for ws in workbook.worksheets:
    trim_sheet(ws)
"""

class Sandbox:    
    def __init__(self, base_dir: Path) -> None:
        """
        Initializes the sandbox environment with a restricted base directory for file operations.
        
        The sandbox provides a secure execution environment for running code on workbooks.
        All file operations are restricted to the specified base directory to prevent
        unauthorized access to the file system.

        Args:
            base_dir: The base directory for all file operations. All file access will be
                     restricted to this directory and its subdirectories.
        """
        self.interpreter = code.InteractiveInterpreter()
        self.code_history = []
        self.stdout = []
        self.stderr = []
        
        self.base_dir = base_dir.resolve()
        if not self.base_dir.is_dir():
            self.base_dir.mkdir(parents=True, exist_ok=True)
            
        self.import_lib()

    def import_lib(self):
        code_import = (
            "import openpyxl\nimport pandas as pd\nimport matplotlib.pyplot as plt\nimport os\nimport datetime\n"
        )

        response = self.step(code_import, dummy=False)
        if response.code == EXEC_CODE.FAIL:
            raise RuntimeError(f"Sandbox: Failed to import essential libraries. Error: {response.msg.strip()}")

    def load_workbook(self, workbook_path):
        path_str = str(workbook_path).replace("\\", "\\\\")

        code_init_wb_path = f'wb_path = r"{path_str}"'
        path_response = self.step(code_init_wb_path, dummy=False)
        if path_response.code == EXEC_CODE.FAIL:
            raise ValueError(f"Sandbox: Failed to set workbook path variable. Error: {path_response.msg.strip()}")

        code_init_load_wb = "workbook = openpyxl.load_workbook(wb_path)"
        
        load_response = self.step(code_init_load_wb, dummy=False)
        if load_response.code == EXEC_CODE.FAIL:
            raise ValueError(f"Sandbox: Failed to load workbook '{workbook_path}'. Error: {load_response.msg.strip()}")

        trim_workbook_code = TRIM_SHEET_CODE
        trim_response = self.step(trim_workbook_code, dummy=False)
        if trim_response.code == EXEC_CODE.FAIL:
            raise ValueError(f"Sandbox: Failed to trim workbook. Error: {trim_response.msg.strip()}")

    def get_existing_sheet_names(self):
        code_snippet = "print(workbook.sheetnames)"
        response = self.step(code_snippet, dummy=True)

        if response.code == EXEC_CODE.FAIL:
            raise ValueError(f"Sandbox: Failed to get sheet names. Underlying interpreter error: {response.msg.strip()}")

        output_str = response.msg
        lines = output_str.strip().splitlines()

        if not lines:
            raise ValueError("No output received from interpreter when getting sheet names (stdout was empty).")

        last_line = lines[-1]

        try:
            sheet_names = ast.literal_eval(last_line)
        except Exception as e:
            raise ValueError(f"Failed to parse sheet names from interpreter output: '{last_line}'. Error: {e}") from e

        if not isinstance(sheet_names, list):
            raise TypeError(f"Expected list of sheet names from interpreter, got {type(sheet_names)}: '{last_line}'")

        return sheet_names

    def get_sheet_state(self) -> str:
        sheet_state = ""
        sheet_names = self.get_existing_sheet_names()
        for sheet_name in sheet_names:
            code_snippet = f"""print(workbook["{sheet_name}"].max_column)
print(workbook["{sheet_name}"].max_row)
print(workbook["{sheet_name}"].min_column)
print(workbook["{sheet_name}"].min_row)
print(workbook["{sheet_name}"].cell(1, 1).value)"""
            max_column, max_row, min_column, min_row, first_value = self.step(code_snippet, dummy=True).msg.splitlines()
            if (max_column == max_row == min_column == min_row == str(1)) and str(first_value) == "None":
                sheet_desc = 'Sheet "{sheet_name}" is empty. '.format(sheet_name=sheet_name)
            else:
                sheet_desc = 'Sheet "{sheet_name}" has {n_rows} rows (Including the header row) and {n_cols} columns ({headers}). '
                code_snippet = f"""print(workbook["{sheet_name}"].max_column)
print(workbook["{sheet_name}"].max_row)
print([cell.value for cell in workbook["{sheet_name}"][1]])
print([str(cell.value.__class__) for cell in workbook["{sheet_name}"][2]])"""

                n_cols, n_rows, headers, data_types = self.step(code_snippet, dummy=True).msg.splitlines()
                import datetime

                headers = eval(headers)
                data_types = eval(data_types)

                headers_str = ", ".join(
                    [
                        f'{chr(65 + i)}({i+1}): "{header}" ({data_type})'
                        for i, (header, data_type) in enumerate(zip(headers, data_types))
                    ]
                )
                sheet_desc = sheet_desc.format(
                    sheet_name=sheet_name,
                    n_cols=int(n_cols),
                    headers=headers_str,
                    n_rows=int(n_rows),
                )

            sheet_state += sheet_desc

        return sheet_state

    def reset(self):
        self.interpreter = code.InteractiveInterpreter()

    def step(self, code_snippet: str, dummy=False) -> SandboxResponse:
        logger.info(f"Executing Python code: {code_snippet}")
        out_buffer = io.StringIO()
        err_buffer = io.StringIO()
        sys.stdout = out_buffer
        sys.stderr = err_buffer

        self.interpreter.runcode(code_snippet)

        output = out_buffer.getvalue()
        error = err_buffer.getvalue()

        if not dummy:
            if error == "":
                self.code_history.append(code_snippet)
            self.stdout.append(output)
            self.stderr.append(error)

        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__

        if error != "":  # error caught
            if not dummy:
                self.reset()
                self.step("\n".join(self.code_history), dummy=True)
            
            return SandboxResponse(EXEC_CODE.FAIL, error)
        
        return SandboxResponse(EXEC_CODE.SUCCESS, output)

    def _check_path(self, path_to_check: Path) -> Path:
        """
        Checks if a given path is within the sandbox's base directory.
        
        This is a critical security function that prevents file operations outside
        the designated sandbox directory, protecting the host system from
        potentially malicious file access.

        Args:
            path_to_check: The path to validate.

        Returns:
            The resolved, absolute path if it is safe.

        Raises:
            PermissionError: If the path is outside the sandbox's base directory.
        """
        resolved_path = path_to_check.resolve()
        if not str(resolved_path).startswith(str(self.base_dir)):
             raise PermissionError("File access outside of the sandbox is not allowed.")
        return resolved_path

    def save(self, save_dir: Path):
        """
        Saves the workbook and session data to the specified directory within the sandbox.
        
        This method validates that the save directory is within the sandbox's base directory
        before performing any file operations, ensuring secure file handling.

        Args:
            save_dir: The absolute path to a directory within the sandbox's base_dir.
            
        Raises:
            PermissionError: If the save directory is outside the sandbox's base directory.
        """
        output_path = self._check_path(save_dir)
        output_path.mkdir(exist_ok=True)
        
        self.step(f'workbook.save(r"{output_path / "workbook_new.xlsx"}")', dummy=False)
        with open(output_path / "code.py", "w", encoding="utf-8") as f:
            f.write("\n\n# ============\n".join(self.code_history))

        with open(output_path / "outputs.txt", "w", encoding="utf-8") as f:
            f.write("\n\n# ============\n".join(self.stdout))

        with open(output_path / "errors.txt", "w", encoding="utf-8") as f:
            f.write("\n\n# ============\n".join(self.stderr))

    def save_temp_workbook(self, save_dir: Path):
        """
        Saves a temporary copy of the workbook to the specified directory within the sandbox.
        
        This method validates that the save directory is within the sandbox's base directory
        before performing any file operations, ensuring secure file handling.

        Args:
            save_dir: The absolute path to a directory within the sandbox's base_dir.
            
        Raises:
            PermissionError: If the save directory is outside the sandbox's base directory.
        """
        output_path = self._check_path(save_dir)
        output_path.mkdir(exist_ok=True)
        self.step(f'workbook.close()\nworkbook.save(r"{output_path / "workbook_temp.xlsx"}")', dummy=True)
